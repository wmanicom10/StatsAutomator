from playwright.sync_api import sync_playwright
import re
import time
from datetime import datetime
import openpyxl
import unicodedata


def normalize_title(title):
    normalized = title.strip()
    normalized = normalized.lower()

    apostrophe_chars = [
        "\u2018",
        "\u2019",
        "\u0027",
        "\u02BC",
        "\u02C8",
        "\u02CA",
        "\u02CB",
        "\u201B",
        "\u2032",
        "\u2035",
        "\u0060",
        "\u00B4",
        "\u1FFE",
        "\u02BB",
    ]
    for char in apostrophe_chars:
        normalized = normalized.replace(char, "'")

    dash_chars = [
        "\u2013",
        "\u2014",
        "\u2212",
        "\u2010",
        "\u2011",
        "\u2012",
        "\u2015",
        "\u2043",
        "\uFE58",
        "\uFE63",
        "\uFF0D",
    ]
    for char in dash_chars:
        normalized = normalized.replace(char, "-")

    normalized = unicodedata.normalize('NFKC', normalized)

    normalized = ''.join(char for char in normalized
                         if unicodedata.category(char)[0] != 'C'
                         or char in '\n\r\t')

    normalized = re.sub(r'\s+', ' ', normalized)
    normalized = normalized.strip()
    return normalized

def scrape_letterboxd_watches(film_url):
    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=['--disable-blink-features=AutomationControlled']
        )

        context = browser.new_context(
            viewport={'width': 1920, 'height': 1080},
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        )

        page = context.new_page()

        try:
            print(f"Loading {film_url}...")
            page.goto(film_url, wait_until='domcontentloaded', timeout=30000)
            page.wait_for_timeout(2000)

            title = "Unknown"
            try:
                title_elem = page.locator('h1.headline-1, h1.filmtitle').first
                if title_elem.count() > 0:
                    title = title_elem.text_content().strip()
            except:
                pass

            watch_count = None
            try:
                watch_link = page.locator('.production-statistic.-watches a.tooltip')
                if watch_link.count() > 0:
                    data_title = watch_link.get_attribute('data-original-title')
                    if data_title:
                        numbers = re.findall(r'([\d,]+)', data_title)
                        if numbers:
                            watch_count = numbers[0]
            except Exception as e:
                print(f"Error extracting watch count: {e}")

            return {
                'title': title,
                'watch_count': watch_count,
                'url': film_url
            }

        except Exception as e:
            print(f"Error: {e}")
            return {
                'title': 'Error',
                'watch_count': None,
                'url': film_url
            }
        finally:
            browser.close()

def scrape_multiple_films(film_urls, delay=5):
    results = []

    for i, url in enumerate(film_urls, 1):
        print(f"\n[{i}/{len(film_urls)}] Scraping...")
        result = scrape_letterboxd_watches(url)
        results.append(result)

        if result['watch_count']:
            print(f"✓ {result['title']}: {result['watch_count']}")
        else:
            print(f"✗ {result['title']}: Could not get watch count")

        if i < len(film_urls):
            print(f"Waiting {delay} seconds...")
            time.sleep(delay)

    return results

def build_movie_column_map(sheet):
    movie_map = {}
    skip_keywords = ['week', 'growth', 'date', 'million', 'premiere', 'theatrical', ' to ']

    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=col).value

        if cell and isinstance(cell, str):
            cell_lower = cell.lower()
            if any(keyword in cell_lower for keyword in skip_keywords):
                continue

            normalized = normalize_title(cell)
            movie_map[normalized] = {
                'column': col,
                'original_name': cell
            }

    return movie_map

def find_date_row(sheet, target_date):
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=1).value
        if isinstance(cell, datetime):
            if cell.date() == target_date.date():
                return row
    return None

def calculate_milestone_date(sheet, movie_col, date_row, milestone_target):
    from datetime import datetime, timedelta

    current_watches = None
    current_row = None
    for row in range(sheet.max_row, 1, -1):
        value = sheet.cell(row=row, column=movie_col).value
        if value and isinstance(value, (int, float)):
            current_watches = value
            current_row = row
            break

    if current_watches is None or current_watches >= milestone_target:
        return None

    weekly_growth_samples = []
    for row in range(3, 47):
        current_val = sheet.cell(row=row, column=movie_col).value
        prev_val = sheet.cell(row=row - 1, column=movie_col).value

        if current_val and prev_val and isinstance(current_val, (int, float)) and isinstance(prev_val, (int, float)):
            weekly_growth = current_val - prev_val
            if weekly_growth > 0:
                weekly_growth_samples.append(weekly_growth)

    if not weekly_growth_samples and current_row:
        for row in range(2, current_row + 1):
            current_val = sheet.cell(row=row, column=movie_col).value
            prev_val = sheet.cell(row=row - 1, column=movie_col).value

            if current_val and prev_val and isinstance(current_val, (int, float)) and isinstance(prev_val,
                                                                                                 (int, float)):
                weekly_growth = current_val - prev_val
                if weekly_growth > 0:
                    weekly_growth_samples.append(weekly_growth)

    if not weekly_growth_samples:
        return None

    average_growth_per_week = sum(weekly_growth_samples) / len(weekly_growth_samples)

    watches_needed = milestone_target - current_watches
    weeks_needed = watches_needed / average_growth_per_week
    days_needed = weeks_needed * 7

    today = datetime.now()
    milestone_date = today + timedelta(days=days_needed)

    return milestone_date

def find_milestone_column(sheet, movie_col):
    for offset in range(1, 10):
        check_col = movie_col + offset
        if check_col > sheet.max_column:
            break

        header = sheet.cell(row=1, column=check_col).value
        if header and isinstance(header, str) and 'Million Date' in header:
            return check_col

    return None

def update_spreadsheet(excel_path, scraped_data, target_date=None):
    if target_date is None:
        target_date = datetime.now()

    print(f"\nOpening {excel_path}...")
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active

    movie_columns = build_movie_column_map(sheet)

    date_row = find_date_row(sheet, target_date)

    if date_row is None:
        print(f"Could not find row for date {target_date.strftime('%Y-%m-%d')}")
        wb.close()
        return

    print(f"✓ Found date row: {date_row} (Date: {target_date.strftime('%Y-%m-%d')})")

    last_week_row = date_row - 1
    updated_count = 0
    not_found_count = 0

    print(f"\n{'=' * 60}")
    print("UPDATES:")
    print(f"{'=' * 60}\n")

    for movie_data in scraped_data:
        title = movie_data['title']
        watch_count = movie_data['watch_count']

        if not watch_count:
            print(f"Skipping {title}: No watch count\n")
            continue

        normalized = normalize_title(title)
        movie_info = movie_columns.get(normalized)

        if movie_info is None:
            print(f"NOT FOUND in spreadsheet")
            print(f"DEBUG - Byte-level analysis:")
            print(f"Normalized bytes: {normalized.encode('utf-8')}")
            print(f"Normalized repr: {repr(normalized)}")
            print(f"Length: {len(normalized)}")

            print(f"\n   Checking similar keys:")
            for key, info in movie_columns.items():
                if 'philosopher' in key or 'harry potter' in key:
                    print(f"   Key: '{key}'")
                    print(f"   Key bytes: {key.encode('utf-8')}")
                    print(f"   Key repr: {repr(key)}")
                    print(f"   Length: {len(key)}")
                    print(f"   Exact match: {key == normalized}")

                    if key != normalized and len(key) == len(normalized):
                        print(f"   Character-by-character comparison:")
                        for i, (c1, c2) in enumerate(zip(key, normalized)):
                            if c1 != c2:
                                print(f"      Position {i}: '{c1}' (U+{ord(c1):04X}) vs '{c2}' (U+{ord(c2):04X})")
            print()
            not_found_count += 1
            continue

        col = movie_info['column']
        original_name = movie_info['original_name']

        print(original_name)

        watch_count_int = int(watch_count.replace(',', ''))
        last_week_value = sheet.cell(row=last_week_row, column=col).value
        last_week_count = last_week_value if isinstance(last_week_value, int) else None

        if last_week_count:
            growth = watch_count_int - last_week_count
            growth_str = f"+{growth:,}" if growth > 0 else f"{growth:,}"
        else:
            growth = None
            growth_str = "N/A"

        milestone_col = find_milestone_column(sheet, col)
        milestone_header = None
        milestone_display = None

        if milestone_col:
            milestone_header = sheet.cell(row=1, column=milestone_col).value

            if '8 Million' in milestone_header:
                milestone_target = 8_000_000
            elif '7 Million' in milestone_header:
                milestone_target = 7_000_000
            elif '5 Million' in milestone_header:
                milestone_target = 5_000_000
            else:
                milestone_target = None

            if milestone_target:
                milestone_date = calculate_milestone_date(sheet, col, date_row, milestone_target)
                if milestone_date:
                    milestone_display = milestone_date.strftime('%Y-%m-%d')
                else:
                    milestone_display = "(already reached or no data)"
            else:
                milestone_display = "(unknown milestone)"

        cell = sheet.cell(row=date_row, column=col)
        cell.value = watch_count_int
        cell.number_format = '#,##0'

        print(f"This week:  {watch_count_int:,}")
        if last_week_count:
            print(f"Last week:  {last_week_count:,}")
            print(f"Growth:     {growth_str} ({growth / 7:,.0f}/day)")
        else:
            print(f"Last week:  (no prior data)")

        print()
        updated_count += 1

    print("Saving changes...")
    wb.save(excel_path)
    wb.close()

    print(f"\n{'=' * 60}")
    print(f"SUMMARY:")
    print(f"  Updated: {updated_count}")
    print(f"  Not found: {not_found_count}")
    print(f"  Skipped: {len(scraped_data) - updated_count - not_found_count}")
    print(f"{'=' * 60}")

if __name__ == "__main__":
    EXCEL_PATH = "/Users/willmanicom/Documents/Misc/movies/movies.xlsx"

    films = [
        "https://letterboxd.com/film/interstellar/",
        "https://letterboxd.com/film/fight-club/",
        "https://letterboxd.com/film/barbie/",

        "https://letterboxd.com/film/walle/",
        "https://letterboxd.com/film/finding-nemo/",
        "https://letterboxd.com/film/kill-bill-vol-1/",
        "https://letterboxd.com/film/the-perks-of-being-a-wallflower/",
        "https://letterboxd.com/film/black-panther/",
        "https://letterboxd.com/film/the-avengers-2012/",
        "https://letterboxd.com/film/the-lord-of-the-rings-the-fellowship-of-the-ring/",
        "https://letterboxd.com/film/django-unchained/",
        "https://letterboxd.com/film/the-shining/",
        "https://letterboxd.com/film/monsters-inc/",
        "https://letterboxd.com/film/the-odyssey-2026/",
        "https://letterboxd.com/film/obsession-2025/",
        "https://letterboxd.com/film/project-hail-mary/",
        "https://letterboxd.com/film/spider-man-brand-new-day/"
    ]

    print("LETTERBOXD WATCH COUNT SCRAPER")
    print("=" * 60)
    print(f"Today's date: {datetime.now().strftime('%-m/%-d/%y')}")
    print(f"Films to scrape: {len(films)}")
    print("=" * 60)

    print("\n" + "=" * 60)
    print("STEP 1: SCRAPING LETTERBOXD")
    print("=" * 60)
    scraped_data = scrape_multiple_films(films, delay=5)

    print("\n" + "=" * 60)
    print("STEP 2: UPDATING SPREADSHEET")
    print("=" * 60)
    update_spreadsheet(EXCEL_PATH, scraped_data)

    print("\nDone!")
